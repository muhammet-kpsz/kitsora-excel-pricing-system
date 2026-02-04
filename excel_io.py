import os
import openpyxl
from openpyxl import Workbook

class ExcelHandler:
    def __init__(self):
        pass

    def get_headers(self, filepath):
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheet = wb.active
            headers = []
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = list(row)
                break
            wb.close()
            return headers
        except Exception as e:
            print(f"Error reading headers: {e}")
            return []

    def get_preview(self, filepath, limit=50):
        # We'll use get_all_rows with a limit for legacy calls, 
        # or just reimplement to be safe.
        return self.get_all_rows(filepath, limit)

    def get_all_rows(self, filepath, limit=None):
        rows = []
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheet = wb.active
            headers = []
            
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    headers = list(row)
                    continue
                
                # Create dict based on headers
                row_data = {}
                for idx, val in enumerate(row):
                    if idx < len(headers):
                        row_data[headers[idx]] = val
                
                rows.append(row_data)
                if limit and len(rows) >= limit:
                    break
            wb.close()
            return rows
        except Exception as e:
            print(f"Error reading rows: {e}")
            return []

    def process_and_save_generator(self, filepath, settings_manager, pricing_engine):
        """
        Generator that yields progress updates:
        (status_type, data)
        status_type: "PART_START", "PROGRESS", "PART_COMPLETE", "DONE", "ERROR"
        """
        # Create logs directory if it doesn't exist
        from datetime import datetime
        
        logs_dir = "logs"
        os.makedirs(logs_dir, exist_ok=True)
        
        # Create timestamped log file
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        log_filename = os.path.join(logs_dir, f"debug_export_{timestamp}.log")
        
        # Open debug log file
        debug_log = open(log_filename, "w", encoding="utf-8")
        
        def log_debug(msg):
            debug_log.write(f"{msg}\n")
            debug_log.flush()
            print(f"[DEBUG] {msg}")
            # Return tuple for yielding to GUI
            return ("LOG", f"{msg}")
        
        try:
            mappings = settings_manager.get("mappings")
            out_config = settings_manager.get("output")
            max_rows = int(out_config.get("max_rows_per_file", 5000))
            out_dir = out_config.get("output_dir", os.path.dirname(filepath))
            filename_template = out_config.get("filename_template", "output_part_{n}.xlsx")
            
            # Target columns to update
            targets = settings_manager.get("targets")
            col_discounted = mappings.get("discounted_price_col")
            col_sell = mappings.get("sell_price_col")
            col_market = mappings.get("market_price_col")
            
            # LOG: Initial settings
            yield log_debug("=" * 80)
            yield log_debug("EXPORT BAŞLADI")
            yield log_debug(f"Kaynak dosya: {filepath}")
            yield log_debug(f"Targets: {targets}")
            yield log_debug(f"Base Price Source: {settings_manager.get('base_price_source', 'NOT SET')}")
            yield log_debug(f"Column mappings:")
            yield log_debug(f"  - Discounted: '{col_discounted}'")
            yield log_debug(f"  - Sell: '{col_sell}'")
            yield log_debug(f"  - Market: '{col_market}'")
            yield log_debug("=" * 80)
            
            # Open Source
            wb_src = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheet_src = wb_src.active
            
            headers = []
            header_map = {} 
            
            row_iterator = sheet_src.iter_rows(values_only=True)
            
            try:
                first_row = next(row_iterator)
                headers = list(first_row)
                yield log_debug(f"\nHeaders okundu ({len(headers)} sütun):")
                for idx, h in enumerate(headers):
                    # Force string for robust matching with settings (UI Combos use strings)
                    h_str = str(h) if h is not None else ""
                    header_map[h_str] = idx
                    yield log_debug(f"  [{idx}] '{h_str}'")
                
                yield log_debug(f"\nHeader map oluşturuldu:")
                yield log_debug(f"  Total headers: {len(header_map)}")
                yield log_debug(f"  Discounted column '{col_discounted}' in map: {col_discounted in header_map}")
                yield log_debug(f"  Sell column '{col_sell}' in map: {col_sell in header_map}")
                yield log_debug(f"  Market column '{col_market}' in map: {col_market in header_map}")
            except StopIteration:
                yield log_debug("HATA: Dosya boş")
                debug_log.close()
                yield ("ERROR", "Dosya boş.")
                return 
            
            # Count total rows for estimation (optional, skip for speed or just count)
            # We will just proceed.
            
            part_num = 1
            current_row_count = 0
            total_processed = 0
            
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.append(headers) 
            
            yield ("PART_START", part_num)
            
            yield log_debug(f"\n{'='*80}")
            yield log_debug("SATIRLAR İŞLENMEYE BAŞLIYOR")
            yield log_debug(f"{'='*80}\n")
            
            row_num = 0
            update_count = {"discounted": 0, "sell": 0, "market": 0}

            for row_vals in row_iterator:
                row_num += 1
                row_vals = list(row_vals)
                
                # Pad row to match headers length (fix for OpenPyXL truncating trailing empty cells)
                if len(row_vals) < len(headers):
                    row_vals.extend([None] * (len(headers) - len(row_vals)))
                
                # Build row dict
                row_dict = {}
                for h, idx in header_map.items():
                    if idx < len(row_vals):
                        row_dict[h] = row_vals[idx]
                
                # Calculate
                res = pricing_engine.calculate_row(row_dict)
                
                # Log first 5 rows in detail
                if row_num <= 5:
                    yield log_debug(f"\n--- Satır {row_num} ---")
                    yield log_debug(f"Hesaplama sonucu:")
                    yield log_debug(f"  final_discounted_price: {res.get('final_discounted_price', 'N/A')}")
                    yield log_debug(f"  label_price: {res.get('label_price', 'N/A')}")
                    yield log_debug(f"  error: {res.get('error', 'YOK')}")
                
                if "error" not in res:
                    # Update Discounted Price
                    if targets.get("update_discounted") and col_discounted:
                        if col_discounted in header_map:
                            idx = header_map[col_discounted]
                            old_value = row_vals[idx]
                            new_value = res["final_discounted_price"]
                            row_vals[idx] = new_value
                            update_count["discounted"] += 1
                            
                            if row_num <= 5:
                                yield log_debug(f"  İndirimli Fiyat güncellendi: {old_value} -> {new_value}")
                            
                            # Verify update
                            if row_vals[idx] != new_value:
                                yield log_debug(f"  UYARI: Satır {row_num}, index {idx} güncellenemedi, zorlanıyor...")
                                row_vals[idx] = new_value
                        else:
                            if row_num <= 5:
                                yield log_debug(f"  UYARI: '{col_discounted}' sütunu header_map'te bulunamadı!")
                    elif row_num <= 5:
                        yield log_debug(f"  İndirimli Fiyat atlandı (target: {targets.get('update_discounted')}, col: '{col_discounted}')")
                    
                    # Update Sell Price
                    if targets.get("update_sell") and col_sell:
                        if col_sell in header_map:
                            idx = header_map[col_sell]
                            old_value = row_vals[idx]
                            new_value = res["label_price"]
                            row_vals[idx] = new_value
                            update_count["sell"] += 1
                            
                            if row_num <= 5:
                                yield log_debug(f"  Satış Fiyatı güncellendi: {old_value} -> {new_value}")
                            
                            # Verify update
                            if row_vals[idx] != new_value:
                                yield log_debug(f"  UYARI: Satır {row_num}, index {idx} güncellenemedi, zorlanıyor...")
                                row_vals[idx] = new_value
                        else:
                            if row_num <= 5:
                                yield log_debug(f"  UYARI: '{col_sell}' sütunu header_map'te bulunamadı!")
                    elif row_num <= 5:
                        yield log_debug(f"  Satış Fiyatı atlandı (target: {targets.get('update_sell')}, col: '{col_sell}')")
                    
                    # Update Market Price
                    if targets.get("update_market") and col_market:
                        if col_market in header_map:
                            idx = header_map[col_market]
                            old_value = row_vals[idx]
                            new_value = res["label_price"]
                            row_vals[idx] = new_value
                            update_count["market"] += 1
                            
                            if row_num <= 5:
                                yield log_debug(f"  Piyasa Fiyatı güncellendi: {old_value} -> {new_value}")
                            
                            # Verify update
                            if row_vals[idx] != new_value:
                                yield log_debug(f"  UYARI: Satır {row_num}, index {idx} güncellenemedi, zorlanıyor...")
                                row_vals[idx] = new_value
                        else:
                            if row_num <= 5:
                                yield log_debug(f"  UYARI: '{col_market}' sütunu header_map'te bulunamadı!")
                    elif row_num <= 5:
                        yield log_debug(f"  Piyasa Fiyatı atlandı (target: {targets.get('update_market')}, col: '{col_market}')")
                
                # ===== NEW FEATURE: Apply export filters (stock + category) =====
                # Get filter settings
                stock_col = mappings.get("stock_col", "")
                include_zero_stock = mappings.get("include_zero_stock", True)
                selected_categories = settings_manager.get("selected_categories", [])
                
                # Stock filter
                if stock_col and not include_zero_stock:
                    try:
                        from stock_filter import StockFilter
                        stock_val = StockFilter.get_stock_value(row_dict, stock_col)
                        if stock_val <= 0:
                            if row_num <= 5:
                                yield log_debug(f"  Satır atlandı: Stok = 0")
                            continue  # Skip zero stock items
                    except:
                        pass  # If stock filter fails, don't block export
                
                # Category tree filter
                if selected_categories:
                    raw_cat_path = res.get("full_category_path", "")
                    # Normalize path (handle spacing differences: "A>B" vs "A > B")
                    if raw_cat_path:
                        full_cat_path = " > ".join([p.strip() for p in str(raw_cat_path).split(">") if p.strip()])
                    else:
                        full_cat_path = ""
                    
                    main_cat = str(res.get("main_category", ""))
                    
                    cat_match = False
                    for selected in selected_categories:
                        if full_cat_path == selected or \
                           main_cat == selected or \
                           full_cat_path.startswith(selected + " >") or \
                           main_cat.startswith(selected + " >"):
                            cat_match = True
                            break
                    
                    if not cat_match:
                        if row_num <= 5:
                            yield log_debug(f"  Satır atlandı: Kategori seçili değil")
                        continue  # Skip if not in selected categories
                # ===== END NEW FEATURE =====
                
                ws_out.append(row_vals)
                current_row_count += 1
                total_processed += 1
                
                if total_processed % 100 == 0:
                    yield ("PROGRESS", (part_num, current_row_count, total_processed))
                
                # Check split
                if current_row_count >= max_rows:
                    # Save current
                    fname = filename_template.replace("{n}", str(part_num))
                    save_path = os.path.join(out_dir, fname)
                    wb_out.save(save_path)
                    
                    yield ("PART_COMPLETE", (part_num, current_row_count))
                    
                    # Reset
                    part_num += 1
                    current_row_count = 0
                    wb_out = Workbook()
                    ws_out = wb_out.active
                    ws_out.append(headers)
                    yield ("PART_START", part_num)
            
            # Save valid leftover
            if current_row_count > 0:
                fname = filename_template.replace("{n}", str(part_num))
                save_path = os.path.join(out_dir, fname)
                wb_out.save(save_path)
                yield ("PART_COMPLETE", (part_num, current_row_count))
            
            wb_src.close()
            
            yield log_debug(f"\n{'='*80}")
            yield log_debug("İŞLEM TAMAMLANDI")
            yield log_debug(f"Toplam satır: {total_processed}")
            yield log_debug(f"Güncelleme sayıları:")
            yield log_debug(f"  - İndirimli Fiyat: {update_count['discounted']}")
            yield log_debug(f"  - Satış Fiyatı: {update_count['sell']}")
            yield log_debug(f"  - Piyasa Fiyatı: {update_count['market']}")
            yield log_debug(f"{'='*80}")
            debug_log.close()
            
            yield ("DONE", f"Toplam {total_processed} satır işlendi, {part_num} dosya oluşturuldu.")
            
        except Exception as e:
            yield log_debug(f"\nFATAL HATA: {str(e)}")
            import traceback
            yield log_debug(traceback.format_exc())
            debug_log.close()
            yield ("ERROR", str(e))
